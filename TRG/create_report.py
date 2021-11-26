import openpyxl
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont

pdfmetrics.registerFont(TTFont('Vera','Vera.ttf'))
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.get_sheet_by_name('Request')

page_width = 2156
page_height = 3050

report_title = 'Technical Report'

print(sheet.cell(2,4).value)

def create_report():
    for i in range(2,42):

        #taking the data from excel sheet
        customer_id = sheet.cell(row = i,column=2).value
        customer_relation = sheet.cell(row = 1,column = 3).value
        customer_service_id = 'something'
        # TODO: add the rest of the data

        #setting up the documents properties
        c = canvas.Canvas(str(customer_id)+'_'+str(customer_id)+'_'+str(customer_service_id)+'.pdf')
        c.setPageSize((page_width,page_height))
        c.setFont('Vera',80)

        text_width = stringWidth(report_title,'Vera',60)
        c.drawString((page_width-page_height)/1000,2900,report_title)

        text = 'This is under the header' + 'date'
        text_width = stringWidth(text,'Vera',55)
        c.setFont('Vera',55)
        c.drawString((page_width-page_height)/1000,2700,report_title)

        c.save()

create_report()
